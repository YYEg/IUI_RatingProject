from openpyxl import load_workbook
from openpyxl.styles import Font
from django.http import FileResponse, HttpResponse
import shutil
import tempfile
from .models import AchievmentGroup, Achievment, EmployeeAchievment, Employee

def generate_personal_report(request):
    # Получаем teacher_id из параметров запроса
    teacher_id = request.GET.get('teacher_id')
    employee = Employee.objects.filter(id=teacher_id).values(
            'surname', 'name', 'parentName', 'position'
        )
    employee = list(employee)

    Name = employee[0]['surname'] + " " + employee[0]['name'] + " " + employee[0]['parentName']
    Dolznost = employee[0]['position']
    
    # Если teacher_id не передан, возвращаем ошибку
    if not teacher_id:
        return HttpResponse("Teacher ID is required", status=400)

    # Путь к файлу шаблона
    file_path = 'personal_example.xlsx'
    sheet_name = 'Лист1'

    # Создаем временный файл на диске (копия шаблона)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        # Копируем шаблон во временный файл
        shutil.copyfile(file_path, tmp_file.name)

        # Загружаем книгу из временного файла
        wb = load_workbook(tmp_file.name)
        ws = wb[sheet_name]

        ws['A7'].value = Name  # Заменяем значение ячейки A6
        ws['A5'].value = Dolznost  # Заменяем значение ячейки A6

        # Получаем и заполняем группы
        groups = AchievmentGroup.objects.all().values('id', 'name', 'description')
        groups = list(groups)
        print("Groups:", groups)  # Отладочный вывод

        # Получаем и заполняем показатели
        achievements = Achievment.objects.all().values('id', 'group_id', 'parent_id', 'number', 'name', 'meas_unit', 'meas_unit_score', 'verif_doc_info')
        achievements = list(achievements)
        print("Achievements:", achievements)  # Отладочный вывод

        # Получаем достижения сотрудника с teacher_id
        employee_achievements = EmployeeAchievment.objects.filter(employee_id=teacher_id).values(
            'id', 'achievment_id', 'full_achivment_name', 'meas_unit_val', 'score'
        )
        employee_achievements = list(employee_achievements)
        print("Employee Achievements:", employee_achievements)  # Отладочный вывод

        # Создаем словарь для суммирования баллов по каждому достижению
        achievements_score = {}

        # Суммируем баллы для каждого достижения
        for emp_ach in employee_achievements:
            if emp_ach["achievment_id"] not in achievements_score:
                achievements_score[emp_ach["achievment_id"]] = 0
            achievements_score[emp_ach["achievment_id"]] += emp_ach["score"]  # Суммируем баллы для достижения

        # Создаем словарь для хранения баллов родительских достижений
        parent_achievements_score = {}

        # Проходим по всем достижениям и суммируем баллы для родительских достижений
        for achievement in achievements:
            if achievement["parent_id"]:  # Если у достижения есть родитель
                if achievement["parent_id"] not in parent_achievements_score:
                    parent_achievements_score[achievement["parent_id"]] = 0
                # Добавляем баллы дочернего достижения к родительскому
                parent_achievements_score[achievement["parent_id"]] += achievements_score.get(achievement["id"], 0)

        # Обновляем баллы для родительских достижений
        for parent_id, score in parent_achievements_score.items():
            if parent_id in achievements_score:
                achievements_score[parent_id] += score
            else:
                achievements_score[parent_id] = score

        # Заполняем группы и показатели, начиная с 11 строки
        row_num = 11  # Начинаем с 11 строки
        for group in groups:
            # Записываем группу
            ws[f"A{row_num}"] = f"Группа {group['id']}"
            ws[f"B{row_num}"] = group["name"]
            ws[f"A{row_num}"].font = Font(bold=True)  # Жирный шрифт для группы
            row_num += 1

            # Записываем показатели для этой группы
            for achievement in achievements:
                if achievement["group_id"] == group["id"]:
                    # Записываем показатель
                    ws[f"B{row_num}"] = achievement["number"] + achievement["name"]
                    ws[f"C{row_num}"] = achievement["meas_unit"]
                    if achievement["meas_unit_score"] != 0:
                        ws[f"D{row_num}"] = achievement["meas_unit_score"]
                    ws[f"E{row_num}"] = achievement["verif_doc_info"]

                    # Применяем жирный шрифт к строке показателя
                    for col in ['B', 'C', 'D', 'E', 'F']:
                        ws[f"{col}{row_num}"].font = Font(bold=True)

                    row_num += 1

                    # Получаем все достижения для этого показателя
                    achievements_for_indicator = [
                        ea for ea in employee_achievements
                        if ea["achievment_id"] == achievement["id"]
                    ]

                    # Записываем достижения для этого показателя
                    total_score = achievements_score.get(achievement["id"], 0)  # Сумма баллов для показателя
                    for ea in achievements_for_indicator:
                        ws[f"B{row_num}"] = ea["full_achivment_name"]  # Наименование достижения
                        ws[f"C{row_num}"] = ea["meas_unit_val"]  # Единица измерения
                        ws[f"F{row_num}"] = ea["score"]  # Количество баллов
                        row_num += 1

                    ws[f"F{row_num - len(achievements_for_indicator) - 1}"] = total_score
                        

        # Сохраняем изменения во временный файл
        wb.save(tmp_file.name)

        # Отладочный вывод: проверяем, что данные добавлены
        print("Данные добавлены в файл")

        # Отправляем файл на скачивание
        response = FileResponse(open(tmp_file.name, 'rb'), as_attachment=True, filename=f"Отчет.xlsx")

    return response